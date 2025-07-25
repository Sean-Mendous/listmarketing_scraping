import os
import shutil
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

def detect_row(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook[sheetname]

    for row in range(1, worksheet.max_row + 2):  # +2で念のため1行余分に見る
        if all(cell.value is None for cell in worksheet[row]):
            return row
    return worksheet.max_row + 1 # 万が一全て埋まっていた場合

def create_workbook(path, sheetname, original_column_map):
    workbook = openpyxl.Workbook()
    workbook.save(path)

    worksheet = workbook.active
    worksheet.title = sheetname

    merged_map = {}
    for section in original_column_map.values():
        merged_map.update(section)

    for key, col_letter in merged_map.items():
        cell = worksheet[f"{col_letter}1"]
        cell.value = key
        cell.fill = PatternFill(fill_type="solid", fgColor="D9D9D9") #gray color

    try:
        status = save_workbook(workbook, path)
        if not status:
            raise RuntimeError("Failed to save workbook")
    except Exception as e:
        raise RuntimeError(f"Failed to save workbook: {e}")

    return workbook, worksheet

def acquisition_workbook(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook[sheetname]
    return workbook, worksheet

def search_into_column(sheet: Worksheet, column: str, search_word: str) -> list:
    matched_rows = []

    for cell in sheet[column]:
        if cell.value and str(cell.value).strip() == search_word:
            matched_rows.append(cell.row)

    return sorted(matched_rows)

def write_by_column(original_path, workbook, sheet, row, column_map, value_map):
    for key, value in value_map.items():
        if key in column_map:
            col_letter = column_map[key]
            cell = sheet[f"{col_letter}{row}"]
            if value is None:
                cell.value = "None"  # None の場合は文字列 "None" に置き換えて書き込む
            elif value not in [""]:
                cell.value = value  # 空白はスキップ
    try:
        status = save_workbook(workbook, original_path)
        if not status:
            raise RuntimeError("Failed to save workbook")
    except Exception as e:
        raise RuntimeError(f"Failed to save workbook: {e}")
    
    return True

def save_workbook(workbook, original_path):
    temp_path = original_path + ".tmp"

    try:
        workbook.save(temp_path)

        # 元ファイルのバックアップ（任意）
        backup_path = original_path + ".bak"
        if os.path.exists(original_path):
            shutil.copy2(original_path, backup_path)

        shutil.move(temp_path, original_path)

        return True

    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return False